"""Parse .xlsx and .csv uploads into validated student rows."""

import csv
import io
from typing import BinaryIO

from app.schemas.admin import StudentRow, RowError

# Required and optional column names (lowercased for matching)
REQUIRED_COLUMNS = {"name", "email"}
OPTIONAL_COLUMNS = {"roll_number", "department", "year_of_study"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

# Common header aliases → canonical name
HEADER_ALIASES = {
    "student name": "name",
    "full name": "name",
    "email address": "email",
    "email id": "email",
    "roll no": "roll_number",
    "roll no.": "roll_number",
    "rollnumber": "roll_number",
    "dept": "department",
    "branch": "department",
    "year": "year_of_study",
    "year of study": "year_of_study",
}


def _normalize_header(raw: str) -> str:
    cleaned = raw.strip().lower().replace("_", " ")
    if cleaned in ALL_COLUMNS:
        return cleaned.replace(" ", "_")
    return HEADER_ALIASES.get(cleaned, cleaned.replace(" ", "_"))


def parse_xlsx(file: BinaryIO) -> tuple[list[StudentRow], list[RowError]]:
    import openpyxl

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    raw_headers = next(rows_iter, None)
    if not raw_headers:
        return [], [RowError(row_number=1, field="file", message="Empty file — no header row")]

    headers = [_normalize_header(str(h)) if h else "" for h in raw_headers]
    return _parse_rows(headers, rows_iter, start_row=2)


def parse_csv(file: BinaryIO) -> tuple[list[StudentRow], list[RowError]]:
    text = file.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    raw_headers = next(reader, None)
    if not raw_headers:
        return [], [RowError(row_number=1, field="file", message="Empty file — no header row")]

    headers = [_normalize_header(h) for h in raw_headers]
    return _parse_rows(headers, (tuple(r) for r in reader), start_row=2)


def _parse_rows(headers: list[str], rows, start_row: int):
    valid: list[StudentRow] = []
    errors: list[RowError] = []

    # Check required columns exist
    header_set = set(headers)
    for req in REQUIRED_COLUMNS:
        if req not in header_set:
            errors.append(RowError(row_number=1, field=req, message=f"Missing required column: {req}"))
    if errors:
        return valid, errors

    col_map = {col: i for i, col in enumerate(headers) if col in ALL_COLUMNS}

    for row_idx, row_values in enumerate(rows, start=start_row):
        # Skip completely empty rows
        if not row_values or all(v is None or str(v).strip() == "" for v in row_values):
            continue

        def cell(col: str) -> str | None:
            idx = col_map.get(col)
            if idx is None or idx >= len(row_values):
                return None
            val = row_values[idx]
            return str(val).strip() if val is not None else None

        name = cell("name")
        email = cell("email")

        if not name:
            errors.append(RowError(row_number=row_idx, field="name", message="Name is required"))
            continue
        if not email:
            errors.append(RowError(row_number=row_idx, field="email", message="Email is required"))
            continue

        # Basic email format check
        if "@" not in email:
            errors.append(RowError(row_number=row_idx, field="email", message=f"Invalid email: {email}"))
            continue

        valid.append(StudentRow(
            row_number=row_idx,
            name=name,
            email=email.lower(),
            roll_number=cell("roll_number"),
            department=cell("department"),
            year_of_study=cell("year_of_study"),
        ))

    return valid, errors


_SAMPLE_ROWS = [
    ("Aarav Sharma", "aarav.sharma@college.edu", "CS2101", "Computer Science", "3rd Year"),
    ("Priya Nair", "priya.nair@college.edu", "EC2205", "Electronics", "2nd Year"),
    ("Rahul Gupta", "rahul.gupta@college.edu", "ME2312", "Mechanical", "3rd Year"),
    ("Sneha Reddy", "sneha.reddy@college.edu", "IT2104", "Information Technology", "4th Year"),
    ("Karthik Iyer", "karthik.iyer@college.edu", "EE2208", "Electrical", "2nd Year"),
    ("Ananya Das", "ananya.das@college.edu", "CS2115", "Computer Science", "1st Year"),
    ("Vikram Singh", "vikram.singh@college.edu", "CE2301", "Civil", "3rd Year"),
    ("Meera Joshi", "meera.joshi@college.edu", "BT2102", "Biotechnology", "2nd Year"),
    ("Arjun Patel", "arjun.patel@college.edu", "CS2220", "Computer Science", "4th Year"),
    ("Divya Menon", "divya.menon@college.edu", "CH2107", "Chemical", "1st Year"),
]


def generate_template_xlsx() -> bytes:
    """Generate an .xlsx template with headers and 10 sample rows."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Name", "Email", "Roll Number", "Department", "Year of Study"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for row in _SAMPLE_ROWS:
        ws.append(row)

    for col, width in [(1, 25), (2, 30), (3, 15), (4, 25), (5, 15)]:
        ws.column_dimensions[chr(64 + col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_template_csv() -> bytes:
    """Generate a .csv template with headers and 10 sample rows."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Email", "Roll Number", "Department", "Year of Study"])
    for row in _SAMPLE_ROWS:
        writer.writerow(row)
    return output.getvalue().encode("utf-8")
